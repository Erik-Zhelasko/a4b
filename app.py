from flask import Flask, render_template, request, redirect, session, flash
import psycopg2
import psycopg2.extras
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = "CHANGE_THIS_SECRET_KEY"   # Change this before submission

# ---------------------------
# DATABASE CONNECTION
# ---------------------------
def get_db():
    return psycopg2.connect(
        dbname="company",
        user="owner",
        host="localhost"
    )

# ---------------------------
# LOGIN REQUIRED DECORATOR
# ---------------------------
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect("/login")
        return f(*args, **kwargs)
    return wrapper

# ---------------------------
# ADMIN ONLY DECORATOR (RBAC)
# ---------------------------
def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect("/login")
        if session.get("role") != "admin":
            return "Access denied (Admin only)", 403
        return f(*args, **kwargs)
    return wrapper

# ---------------------------
# LOGIN ROUTE (A1)
# ---------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        cur.execute("SELECT * FROM app_user WHERE username=%s", (username,))
        user = cur.fetchone()

        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            session["role"] = user["role"]
            return redirect("/")
        else:
            flash("Invalid username or password")

    return render_template("login.html")

# ---------------------------
# LOGOUT (A1)
# ---------------------------
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

# ---------------------------
# HOME ROUTE (A2)
# ---------------------------
@app.route("/")
@login_required
def home():
    search = request.args.get("search", "")
    sort = request.args.get("sort", "name_asc")
    dept = request.args.get("dept", "")   # NEW: selected department

    # Sorting whitelist
    allowed_sorts = {
        "name_asc": "e.lname ASC, e.fname ASC",
        "name_desc": "e.lname DESC, e.fname DESC",
        "hours_asc": "total_hours ASC",
        "hours_desc": "total_hours DESC"
    }
    order_by = allowed_sorts.get(sort, "e.lname ASC")

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # NEW: Load department names for dropdown
    cur.execute("SELECT dname FROM department ORDER BY dname;")
    departments = [row[0] for row in cur.fetchall()]

    # Build WHERE clause
    where_clauses = []
    params = []

    # Name search (case-insensitive, partial)
    where_clauses.append("(e.fname ILIKE %s OR e.lname ILIKE %s)")
    params.extend([f"%{search}%", f"%{search}%"])

    # Department filter (optional)
    if dept != "":
        where_clauses.append("d.dname = %s")
        params.append(dept)

    where_sql = " AND ".join(where_clauses)

    # FINAL QUERY
    query = f"""
        SELECT
            e.ssn,
            e.fname || ' ' || e.lname AS full_name,
            d.dname,
            COALESCE(dep.count_dep, 0) AS num_dependents,
            COALESCE(w.count_proj, 0) AS num_projects,
            COALESCE(w.total_hours, 0) AS total_hours
        FROM employee e
        JOIN department d ON e.dno = d.dnumber

        LEFT JOIN (
            SELECT essn, COUNT(*) AS count_dep
            FROM dependent
            GROUP BY essn
        ) dep ON dep.essn = e.ssn

        LEFT JOIN (
            SELECT essn, COUNT(*) AS count_proj, SUM(hours) AS total_hours
            FROM works_on
            GROUP BY essn
        ) w ON w.essn = e.ssn

        WHERE {where_sql}
        ORDER BY {order_by};
    """

    cur.execute(query, tuple(params))
    employees = cur.fetchall()

    return render_template(
        "home.html",
        employees=employees,
        departments=departments,   # NEW
        selected_dept=dept,        # NEW
        search=search              # so form remembers search
    )

# ---------------------------
# PROJECTS ROUTE (A3)
# ---------------------------

@app.route("/projects")
@login_required
def projects():

    # Sorting whitelist
    sort = request.args.get("sort", "headcount_desc")
    allowed_sorts = {
        "headcount_asc": "headcount ASC",
        "headcount_desc": "headcount DESC",
        "hours_asc": "total_hours ASC",
        "hours_desc": "total_hours DESC"
    }
    order_by = allowed_sorts.get(sort, "headcount DESC")

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Query for project summary
    query = f"""
        SELECT 
            p.pnumber,
            p.pname,
            d.dname AS department_name,
            COALESCE(w.headcount, 0) AS headcount,
            COALESCE(w.total_hours, 0) AS total_hours
        FROM project p
        JOIN department d ON p.dnum = d.dnumber

        LEFT JOIN (
            SELECT 
                pno,
                COUNT(DISTINCT essn) AS headcount,
                SUM(hours) AS total_hours
            FROM works_on
            GROUP BY pno
        ) w ON w.pno = p.pnumber

        ORDER BY {order_by};
    """

    cur.execute(query)
    projects = cur.fetchall()

    return render_template(
        "projects.html",
        projects=projects,
        sort=sort
    )

# ---------------------------
# PROJECT DETAIL ROUTE (A4)
# ---------------------------
@app.route("/project/<int:pnum>", methods=["GET", "POST"])
@login_required
@admin_required
def project_detail(pnum):

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Handle form submission (assign/update hours)
    if request.method == "POST":
        essn = request.form["essn"]
        hours = request.form["hours"]

        # Upsert works_on (insert or update)
        cur.execute("""
            INSERT INTO works_on (essn, pno, hours)
            VALUES (%s, %s, %s)
            ON CONFLICT (essn, pno)
            DO UPDATE SET hours = EXCLUDED.hours;
        """, (essn, pnum, hours))

        conn.commit()

        return redirect(f"/project/{pnum}")

    # Load project info
    cur.execute("""
        SELECT p.pnumber, p.pname, d.dname
        FROM project p
        JOIN department d ON p.dnum = d.dnumber
        WHERE p.pnumber = %s;
    """, (pnum,))
    project = cur.fetchone()

    if not project:
        return "Project not found", 404

    # Load employees assigned to this project
    cur.execute("""
        SELECT
            e.ssn,
            e.fname || ' ' || e.lname AS full_name,
            COALESCE(w.hours, 0) AS hours
        FROM employee e
        LEFT JOIN works_on w ON w.essn = e.ssn AND w.pno = %s
        ORDER BY e.lname, e.fname;
    """, (pnum,))
    employees = cur.fetchall()

    return render_template(
        "project_details.html",
        project=project,
        employees=employees
    )

# ---------------------------
# EMPLOYEE DETAIL ROUTE (A5)
# ---------------------------
@app.route("/employee/<string:ssn>", methods=["GET", "POST"])
@login_required
@admin_required
def employee_management(ssn):

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Handle add/update dependent
    if request.method == "POST":
        dname = request.form["dname"]
        sex = request.form["sex"]
        bdate = request.form["bdate"]
        relationship = request.form["relationship"]

        cur.execute("""
            INSERT INTO dependent (essn, dependent_name, sex, bdate, relationship)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (essn, dependent_name)
            DO UPDATE SET
                sex = EXCLUDED.sex,
                bdate = EXCLUDED.bdate,
                relationship = EXCLUDED.relationship;
        """, (ssn, dname, sex, bdate, relationship))

        conn.commit()
        return redirect(f"/employee/{ssn}")

    # Load employee info
    cur.execute("""
        SELECT ssn, fname || ' ' || lname AS full_name, address, salary, dno
        FROM employee
        WHERE ssn = %s;
    """, (ssn,))
    employee = cur.fetchone()

    if not employee:
        return "Employee not found", 404

    # Load dependents
    cur.execute("""
        SELECT dependent_name, sex, bdate, relationship
        FROM dependent
        WHERE essn = %s
        ORDER BY dependent_name;
    """, (ssn,))
    dependents = cur.fetchall()

    return render_template(
        "employee_management.html",
        employee=employee,
        dependents=dependents
    )

# ---------------------------
# MANAGER OVERVIEW (A6)
# ---------------------------
@app.route("/manager_overview")
@login_required
def manager_overview():

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query = """
        SELECT
            d.dname,
            d.dnumber,
            COALESCE(m.fname || ' ' || m.lname, 'N/A') AS manager_name,
            COALESCE(emp.count_emp, 0) AS num_employees,
            COALESCE(hours.total_hours, 0) AS total_hours
        FROM department d

        LEFT JOIN employee m
            ON d.mgr_ssn = m.ssn

        LEFT JOIN (
            SELECT dno, COUNT(*) AS count_emp
            FROM employee
            GROUP BY dno
        ) emp
            ON emp.dno = d.dnumber

        LEFT JOIN (
            SELECT e.dno AS dept_no,
                   COALESCE(SUM(w.hours), 0) AS total_hours
            FROM employee e
            LEFT JOIN works_on w ON e.ssn = w.essn
            GROUP BY e.dno
        ) hours
            ON hours.dept_no = d.dnumber

        ORDER BY d.dnumber;
    """

    cur.execute(query)
    overview = cur.fetchall()

    return render_template("manager_overview.html", overview=overview)

# ---------------------------
# CSV EXPORT
# ---------------------------
@app.route("/export_employees")
@login_required
def export_employees():

    search = request.args.get("search", "")
    dept = request.args.get("dept", "")
    sort = request.args.get("sort", "name_asc")

    allowed_sorts = {
        "name_asc": "e.lname ASC, e.fname ASC",
        "name_desc": "e.lname DESC, e.fname DESC",
        "hours_asc": "total_hours ASC",
        "hours_desc": "total_hours DESC"
    }
    order_by = allowed_sorts.get(sort, "e.lname ASC")

    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    where_clauses = ["(e.fname ILIKE %s OR e.lname ILIKE %s)"]
    params = [f"%{search}%", f"%{search}%"]

    if dept != "":
        where_clauses.append("d.dname = %s")
        params.append(dept)

    where_sql = " AND ".join(where_clauses)

    query = f"""
        SELECT 
            e.fname || ' ' || e.lname AS full_name,
            d.dname,
            COALESCE(dep.count_dep,0) AS num_dependents,
            COALESCE(w.count_proj,0) AS num_projects,
            COALESCE(w.total_hours,0) AS total_hours
        FROM employee e
        JOIN department d ON e.dno = d.dnumber
        LEFT JOIN (
            SELECT essn, COUNT(*) AS count_dep
            FROM dependent
            GROUP BY essn
        ) dep ON dep.essn = e.ssn
        LEFT JOIN (
            SELECT essn, COUNT(*) AS count_proj, SUM(hours) AS total_hours
            FROM works_on
            GROUP BY essn
        ) w ON w.essn = e.ssn
        WHERE {where_sql}
        ORDER BY {order_by};
    """

    cur.execute(query, tuple(params))
    rows = cur.fetchall()

    # --- Build CSV ---
    import csv
    from io import StringIO
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Name","Department","Dependents","Projects","Total Hours"])

    for r in rows:
        writer.writerow([r["full_name"], r["dname"], r["num_dependents"], r["num_projects"], r["total_hours"]])

    return (
        output.getvalue(),
        200,
        {
            "Content-Type": "text/csv",
            "Content-Disposition": "attachment; filename=employees_export.csv"
        }
    )

# ---------------------------
# EXCEL IMPORT
# ---------------------------
@app.route("/import_dependents", methods=["POST"])
@login_required
@admin_required
def import_dependents():

    from openpyxl import load_workbook
    from io import BytesIO

    file = request.files.get("file")
    if not file or not file.filename.endswith(".xlsx"):
        flash("Please upload a valid .xlsx file.")
        return redirect(request.referrer)

    try:
        wb = load_workbook(BytesIO(file.read()))
        sheet = wb.active

        rows = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                continue  # skip header

            essn, name, sex, bdate, relationship = row

            # Basic validation
            if not essn or not name:
                flash(f"Invalid row {i+1}: ESSN and NAME required.")
                return redirect(request.referrer)

            rows.append((essn, name, sex, bdate, relationship))

        # All rows valid → insert all at once (transaction)
        conn = get_db()
        cur = conn.cursor()

        for r in rows:
            cur.execute("""
                INSERT INTO dependent (essn, dependent_name, sex, bdate, relationship)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (essn, dependent_name)
                DO UPDATE SET sex=EXCLUDED.sex, bdate=EXCLUDED.bdate, relationship=EXCLUDED.relationship;
            """, r)

        conn.commit()
        flash("File imported successfully!")

    except Exception as e:
        flash(f"Import failed: {str(e)}")

    return redirect(request.referrer)


# ---------------------------
# RUN FLASK APP
# ---------------------------
if __name__ == "__main__":
    app.run(debug=True)
